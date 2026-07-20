"""Generate photo-relevance-checklist.xlsx from peaks.json."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
PEAKS_PATH = ROOT / "src" / "data" / "peaks.json"
OUT_PATH = ROOT / "photo-relevance-checklist.xlsx"


def main() -> None:
    peaks = json.loads(PEAKS_PATH.read_text(encoding="utf-8"))
    peaks = sorted(peaks, key=lambda p: (p.get("name") or p["id"]).lower())
    total_photos = sum(len(p.get("photos") or []) for p in peaks)

    wb = Workbook()

    # --- Checklist (first sheet) ---
    ws = wb.active
    ws.title = "Checklist"
    headers = [
        "#",
        "Peak ID",
        "Peak Name",
        "Country",
        "Range",
        "Photo #",
        "Credit",
        "License",
        "Image URL",
        "Source URL",
        "Site page",
        "Status",
        "Notes",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    link_font = Font(color="0563C1", underline="single")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_n = 2
    photo_i = 0
    for peak in peaks:
        photos = peak.get("photos") or []
        if not photos:
            photos = [{}]
        for idx, photo in enumerate(photos, 1):
            photo_i += 1
            peak_id = peak.get("id") or ""
            image_url = photo.get("url") or ""
            source_url = photo.get("sourceUrl") or ""
            site_url = f"https://peakatlas3d.com/peak/{peak_id}" if peak_id else ""
            values = [
                photo_i,
                peak_id,
                peak.get("name") or "",
                peak.get("country") or "",
                peak.get("range") or "",
                idx,
                (photo.get("credit") or "")[:80],
                photo.get("license") or "",
                image_url,
                source_url,
                site_url,
                "",
                "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_n, col, val)
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=(col in (7, 13)))
            if image_url:
                ws.cell(row_n, 9).hyperlink = image_url
                ws.cell(row_n, 9).font = link_font
            if source_url:
                ws.cell(row_n, 10).hyperlink = source_url
                ws.cell(row_n, 10).font = link_font
            if site_url:
                ws.cell(row_n, 11).hyperlink = site_url
                ws.cell(row_n, 11).font = link_font
            if row_n % 2 == 0:
                for col in range(1, 14):
                    if col not in (9, 10, 11):
                        ws.cell(row_n, col).fill = alt_fill
            row_n += 1

    last_row = row_n - 1

    dv = DataValidation(
        type="list",
        formula1='"OK,Replace,Unsure"',
        allow_blank=True,
    )
    dv.error = "Pick OK, Replace, or Unsure"
    dv.errorTitle = "Invalid status"
    dv.prompt = "Is this photo relevant to the peak?"
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(f"L2:L{last_row}")

    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        FormulaRule(formula=['$L2="OK"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        FormulaRule(formula=['$L2="Replace"'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        FormulaRule(formula=['$L2="Unsure"'], fill=PatternFill("solid", fgColor="FFEB9C")),
    )

    widths = {
        "A": 6,
        "B": 14,
        "C": 22,
        "D": 14,
        "E": 22,
        "F": 8,
        "G": 22,
        "H": 14,
        "I": 40,
        "J": 40,
        "K": 36,
        "L": 12,
        "M": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{last_row}"
    ws.row_dimensions[1].height = 28

    # --- Summary ---
    ws_s = wb.create_sheet("Summary", 1)
    ws_s["A1"] = "Progress"
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s["A3"] = "Metric"
    ws_s["B3"] = "Count"
    ws_s["A3"].font = Font(bold=True)
    ws_s["B3"].font = Font(bold=True)
    rows = [
        (4, "Total photos", f"=COUNTA(Checklist!A2:A{last_row})"),
        (5, "Reviewed (any status)", f"=COUNTA(Checklist!L2:L{last_row})"),
        (6, "OK", f'=COUNTIF(Checklist!L2:L{last_row},"OK")'),
        (7, "Replace", f'=COUNTIF(Checklist!L2:L{last_row},"Replace")'),
        (8, "Unsure", f'=COUNTIF(Checklist!L2:L{last_row},"Unsure")'),
        (9, "Still blank", f"=COUNTBLANK(Checklist!L2:L{last_row})"),
    ]
    for r, label, formula in rows:
        ws_s[f"A{r}"] = label
        ws_s[f"B{r}"] = formula
    ws_s["A11"] = "Filter Checklist by Status = (blanks) to see what is left."
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 12

    # --- Instructions ---
    ws_i = wb.create_sheet("Instructions", 2)
    ws_i["A1"] = "PeakAtlas3D — Photo relevance checklist"
    ws_i["A1"].font = Font(bold=True, size=16)
    ws_i["A3"] = "How to use"
    ws_i["A3"].font = Font(bold=True, size=12)
    instructions = [
        "1. Open the Checklist sheet. Filters are on — filter Status blanks to see unreviewed photos.",
        "2. Click Image URL or Source URL to open the photo (Commons source page often has better context).",
        "3. Optionally open Site page to see how it appears on peakatlas3d.com.",
        '4. Set Status: OK (clear scenic shot of this peak) / Replace (wrong peak, cairn-only, map, etc.) / Unsure.',
        "5. Add a short note when Replace or Unsure (e.g. cairn close-up, wrong mountain, too dark).",
        "6. Progress counts update automatically on the Summary sheet.",
        "",
        f"Generated from peaks.json — {len(peaks)} peaks, {total_photos} photos.",
        "Gannett Peak currently has only 1 photo.",
    ]
    for i, line in enumerate(instructions, start=4):
        ws_i[f"A{i}"] = line
    ws_i.column_dimensions["A"].width = 110

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"{len(peaks)} peaks, {total_photos} photo rows")


if __name__ == "__main__":
    main()
